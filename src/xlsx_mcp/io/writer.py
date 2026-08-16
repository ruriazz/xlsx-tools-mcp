from __future__ import annotations

import os
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import range_boundaries
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..errors import SheetNotFoundError
from ..recalc import recalculate


def _load(path: str) -> Workbook:
    return openpyxl.load_workbook(path, data_only=False)


def _ws(wb: Workbook, sheet: str) -> Worksheet:
    if sheet not in wb.sheetnames:
        raise SheetNotFoundError(f"Sheet '{sheet}' not found. Available: {wb.sheetnames}")
    return wb[sheet]


def _atomic_save(wb: Workbook, path: str) -> None:
    """Save to a temp file in the same directory, then atomically replace `path`.

    A plain `wb.save(path)` writes straight onto the target — if the process dies
    mid-write (crash, OOM kill, disk full), the file is left half-written and
    unrecoverable. `os.replace` only swaps the two once the temp file is complete.
    """
    tmp_path = f"{path}.tmp-{os.getpid()}"
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, path)
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def _finalize(wb: Workbook, path: str, recalc: bool = True) -> dict[str, Any]:
    """Save via openpyxl (preserves everything it didn't touch), then recalculate.

    `recalc=False` skips the LibreOffice round-trip for structural edits (merge,
    style, sheet add/remove) that can't produce a stale formula result — recalc is
    still cheap to run for anything that touches cell values or formulas.
    """
    _atomic_save(wb, path)
    wb.close()
    if not recalc:
        return {"saved": True, "recalculated": False, "errors_found": [], "message": "Saved (no recalculation needed)."}

    result = recalculate(path)
    return {
        "saved": True,
        "recalculated": result.success,
        "errors_found": result.errors_found,
        "message": result.message,
    }


_EXCEL_EXTENSIONS = (".xlsx", ".xlsm")


def create_workbook(path: str, sheets: list[str] | None = None, overwrite: bool = False) -> dict[str, Any]:
    p = Path(path)
    if not p.suffix.lower() in _EXCEL_EXTENSIONS:
        raise ValueError(f"Refusing to write to a non-Excel file: {path}. Use a .xlsx/.xlsm path.")
    if p.exists() and not overwrite:
        raise FileExistsError(f"File already exists: {path}. Pass overwrite=True to replace it.")

    wb = openpyxl.Workbook()
    names = sheets or ["Sheet1"]
    wb.active.title = names[0]
    for name in names[1:]:
        wb.create_sheet(name)

    p.parent.mkdir(parents=True, exist_ok=True)
    _atomic_save(wb, path)
    wb.close()
    return {
        "saved": True,
        "path": str(p),
        "sheets": names,
        "recalculated": False,
        "errors_found": [],
        "message": "Workbook created.",
    }


def write_cells(
    path: str,
    sheet: str,
    cells: list[dict[str, Any]],
    create_sheet_if_missing: bool = False,
    recalculate: bool = True,
) -> dict[str, Any]:
    if not cells:
        return {"saved": True, "recalculated": False, "errors_found": [], "message": "Nothing to write; file unchanged."}
    wb = _load(path)
    if sheet not in wb.sheetnames and create_sheet_if_missing:
        wb.create_sheet(sheet)
    ws = _ws(wb, sheet)

    for item in cells:
        coord = item.get("cell")
        if not coord:
            raise ValueError(f"Each item in `cells` requires a 'cell' key, got: {item}")
        formula = item.get("formula")
        ws[coord] = formula if formula is not None else item.get("value")

    return _finalize(wb, path, recalc=recalculate)


def append_rows(
    path: str,
    sheet: str,
    rows: list[list[Any]],
    create_sheet_if_missing: bool = False,
    recalculate: bool = True,
) -> dict[str, Any]:
    if not rows:
        return {"saved": True, "recalculated": False, "errors_found": [], "message": "Nothing to write; file unchanged."}
    wb = _load(path)
    if sheet not in wb.sheetnames and create_sheet_if_missing:
        wb.create_sheet(sheet)
    ws = _ws(wb, sheet)

    for row in rows:
        ws.append(row)

    return _finalize(wb, path, recalc=recalculate)


def create_sheet(path: str, sheet: str, index: int | None = None) -> dict[str, Any]:
    wb = _load(path)
    if sheet in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet}' already exists")
    wb.create_sheet(sheet, index)
    return _finalize(wb, path, recalc=False)


def delete_sheet(path: str, sheet: str) -> dict[str, Any]:
    wb = _load(path)
    _ws(wb, sheet)
    if len(wb.sheetnames) == 1:
        raise ValueError("Cannot delete the only sheet in a workbook")
    del wb[sheet]
    return _finalize(wb, path, recalc=False)


def insert_rows(
    path: str, sheet: str, start_row: int, count: int = 1, recalculate: bool = True
) -> dict[str, Any]:
    wb = _load(path)
    _ws(wb, sheet).insert_rows(start_row, count)
    return _finalize(wb, path, recalc=recalculate)


def delete_rows(
    path: str, sheet: str, start_row: int, count: int = 1, recalculate: bool = True
) -> dict[str, Any]:
    wb = _load(path)
    _ws(wb, sheet).delete_rows(start_row, count)
    return _finalize(wb, path, recalc=recalculate)


def insert_columns(
    path: str, sheet: str, start_column: int, count: int = 1, recalculate: bool = True
) -> dict[str, Any]:
    wb = _load(path)
    _ws(wb, sheet).insert_cols(start_column, count)
    return _finalize(wb, path, recalc=recalculate)


def delete_columns(
    path: str, sheet: str, start_column: int, count: int = 1, recalculate: bool = True
) -> dict[str, Any]:
    wb = _load(path)
    _ws(wb, sheet).delete_cols(start_column, count)
    return _finalize(wb, path, recalc=recalculate)


def merge_cells(path: str, sheet: str, cell_range: str) -> dict[str, Any]:
    wb = _load(path)
    _ws(wb, sheet).merge_cells(cell_range)
    return _finalize(wb, path, recalc=False)


def unmerge_cells(path: str, sheet: str, cell_range: str) -> dict[str, Any]:
    wb = _load(path)
    _ws(wb, sheet).unmerge_cells(cell_range)
    return _finalize(wb, path, recalc=False)


def set_cell_style(path: str, sheet: str, cell_range: str, style: dict[str, Any]) -> dict[str, Any]:
    wb = _load(path)
    ws = _ws(wb, sheet)
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)

    font_kwargs = {k: style[k] for k in ("bold", "italic") if k in style}
    if "font_size" in style:
        font_kwargs["size"] = style["font_size"]
    if "font_color" in style:
        font_kwargs["color"] = style["font_color"]
    font = Font(**font_kwargs) if font_kwargs else None

    fill = None
    if "bg_color" in style:
        fill = PatternFill(start_color=style["bg_color"], end_color=style["bg_color"], fill_type="solid")

    alignment = None
    if "horizontal" in style or "vertical" in style:
        alignment = Alignment(horizontal=style.get("horizontal"), vertical=style.get("vertical"))

    border = None
    if "border" in style:
        side = Side(style=style["border"])
        border = Border(left=side, right=side, top=side, bottom=side)

    number_format = style.get("number_format")

    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            if alignment:
                cell.alignment = alignment
            if border:
                cell.border = border
            if number_format:
                cell.number_format = number_format

    return _finalize(wb, path, recalc=False)
