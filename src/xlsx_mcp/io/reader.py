from __future__ import annotations

import logging
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries
from python_calamine import CalamineWorkbook

from ..errors import SheetNotFoundError

logger = logging.getLogger(__name__)


def list_sheets(path: str) -> list[dict[str, Any]]:
    """List sheet names with their approximate used-range size (calamine, fast)."""
    wb = CalamineWorkbook.from_path(path)
    return [
        {"name": name, "rows": sh.height, "columns": sh.width}
        for name in wb.sheet_names
        for sh in (wb.get_sheet_by_name(name),)
    ]


def workbook_info(path: str) -> dict[str, Any]:
    """Workbook-level metadata that calamine doesn't expose: exact dims, defined names."""
    wb = openpyxl.load_workbook(path, data_only=False)
    try:
        sheets = [
            {
                "name": ws.title,
                "dimensions": ws.dimensions,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "sheet_state": ws.sheet_state,
            }
            for ws in wb.worksheets
        ]
        return {
            "sheets": sheets,
            "active_sheet": wb.active.title if wb.active is not None else None,
            "defined_names": list(wb.defined_names.keys()),
        }
    finally:
        wb.close()


def _read_sheet_openpyxl_values(path: str, sheet: str) -> list[list[Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet}' not found. Available: {wb.sheetnames}")
        return [list(row) for row in wb[sheet].iter_rows(values_only=True)]
    finally:
        wb.close()


def _slice_range(data: list[list[Any]], cell_range: str) -> list[list[Any]]:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    sliced: list[list[Any]] = []
    for r in range(min_row, max_row + 1):
        row = data[r - 1] if r - 1 < len(data) else []
        sliced.append([row[c - 1] if c - 1 < len(row) else None for c in range(min_col, max_col + 1)])
    return sliced


def read_sheet(path: str, sheet: str, cell_range: str | None = None, max_rows: int | None = None) -> dict[str, Any]:
    """Read cell values as a 2D array, addressed absolutely from A1 (row/col 1 = A1).

    Primary path: python-calamine (fast, accurate type inference). Falls back to
    openpyxl if calamine can't parse the file. `skip_empty_area` is disabled so
    coordinates always line up with A1 addressing — required for `cell_range` to
    slice the correct cells.

    `max_rows` optionally caps the number of rows returned, to bound response size
    for large sheets. `None` means no limit.
    """
    try:
        wb = CalamineWorkbook.from_path(path)
        if sheet not in wb.sheet_names:
            raise SheetNotFoundError(f"Sheet '{sheet}' not found. Available: {wb.sheet_names}")
        data = wb.get_sheet_by_name(sheet).to_python(skip_empty_area=False)
    except SheetNotFoundError:
        raise
    except Exception:
        logger.debug("calamine failed to read %s, falling back to openpyxl", path, exc_info=True)
        data = _read_sheet_openpyxl_values(path, sheet)

    if cell_range:
        data = _slice_range(data, cell_range)

    if max_rows is not None:
        data = data[:max_rows]

    return {
        "sheet": sheet,
        "cell_range": cell_range,
        "rows": data,
        "row_count": len(data),
        "column_count": max((len(row) for row in data), default=0),
    }


def _color_hex(color: Any) -> str | None:
    rgb = getattr(color, "rgb", None)
    return rgb if isinstance(rgb, str) else None


def get_cell(path: str, sheet: str, cell: str) -> dict[str, Any]:
    """Full detail for one cell: value, formula, format, font, fill, merge, comment.

    Loads the workbook with `data_only=False` for formulas/styles/comments. The
    `data_only=True` load (for the cached computed value) is only done lazily when
    the cell actually holds a formula, since openpyxl can't yield both from one load.
    """
    wb_formula = openpyxl.load_workbook(path, data_only=False)
    wb_value = None
    try:
        if sheet not in wb_formula.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet}' not found. Available: {wb_formula.sheetnames}")

        ws_formula = wb_formula[sheet]
        c_formula = ws_formula[cell]
        is_formula = isinstance(c_formula.value, str) and c_formula.value.startswith("=")

        if is_formula:
            wb_value = openpyxl.load_workbook(path, data_only=True)
            if sheet not in wb_value.sheetnames:
                raise SheetNotFoundError(f"Sheet '{sheet}' not found. Available: {wb_value.sheetnames}")
            value = wb_value[sheet][cell].value
        else:
            value = c_formula.value

        is_merged = any(cell in mr for mr in ws_formula.merged_cells.ranges)

        return {
            "cell": cell,
            "value": value,
            "formula": c_formula.value if is_formula else None,
            "data_type": c_formula.data_type,
            "number_format": c_formula.number_format,
            "font": {
                "bold": c_formula.font.bold,
                "italic": c_formula.font.italic,
                "size": c_formula.font.size,
                "color": _color_hex(c_formula.font.color),
            },
            "fill_color": _color_hex(c_formula.fill.fgColor) if c_formula.fill else None,
            "is_merged": is_merged,
            "comment": c_formula.comment.text if c_formula.comment else None,
        }
    finally:
        wb_formula.close()
        if wb_value is not None:
            wb_value.close()


def _find_matches_in_sheet(name: str, data: list[list[Any]], needle: str, match_case: bool) -> list[dict[str, Any]]:
    matches: list[dict[str, Any]] = []
    for r_idx, row in enumerate(data, start=1):
        for c_idx, value in enumerate(row, start=1):
            if value is None:
                continue
            haystack = str(value) if match_case else str(value).lower()
            if needle in haystack:
                matches.append({"sheet": name, "cell": f"{get_column_letter(c_idx)}{r_idx}", "value": value})
    return matches


def search_workbook(
    path: str, query: str, sheet: str | None = None, match_case: bool = False, limit: int | None = None
) -> list[dict[str, Any]]:
    """Search cell values across one or all sheets for a substring match (calamine).

    `limit` optionally caps the number of matches returned; `None` means no limit.
    """
    wb = CalamineWorkbook.from_path(path)
    target_sheets = [sheet] if sheet else list(wb.sheet_names)
    needle = query if match_case else query.lower()

    matches: list[dict[str, Any]] = []
    for name in target_sheets:
        if name not in wb.sheet_names:
            raise SheetNotFoundError(f"Sheet '{name}' not found. Available: {wb.sheet_names}")
        data = wb.get_sheet_by_name(name).to_python(skip_empty_area=False)
        matches.extend(_find_matches_in_sheet(name, data, needle, match_case))

    if limit is not None:
        matches = matches[:limit]
    return matches
