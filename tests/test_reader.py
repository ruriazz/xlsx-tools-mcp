import openpyxl
import pytest

from xlsx_mcp.errors import SheetNotFoundError
from xlsx_mcp.io import reader


@pytest.fixture
def workbook_path(tmp_path):
    path = tmp_path / "book.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["item", "qty"])
    ws.append(["apple", 10])
    ws.append(["banana", 5])
    wb.save(path)
    return str(path)


def test_read_sheet_returns_absolute_grid(workbook_path):
    result = reader.read_sheet(workbook_path, "Data")
    assert result["rows"] == [["item", "qty"], ["apple", 10], ["banana", 5]]
    assert result["row_count"] == 3
    assert result["column_count"] == 2


def test_read_sheet_slices_cell_range(workbook_path):
    result = reader.read_sheet(workbook_path, "Data", cell_range="A2:A3")
    assert result["rows"] == [["apple"], ["banana"]]


def test_read_sheet_max_rows(workbook_path):
    result = reader.read_sheet(workbook_path, "Data", max_rows=2)
    assert result["rows"] == [["item", "qty"], ["apple", 10]]
    assert result["row_count"] == 2
    assert reader.read_sheet(workbook_path, "Data")["row_count"] == 3


def test_read_sheet_missing_sheet_raises(workbook_path):
    with pytest.raises(SheetNotFoundError):
        reader.read_sheet(workbook_path, "NoSuchSheet")


def test_list_sheets(workbook_path):
    sheets = reader.list_sheets(workbook_path)
    assert sheets == [{"name": "Data", "rows": 3, "columns": 2}]


def test_get_cell_plain_value(workbook_path):
    info = reader.get_cell(workbook_path, "Data", "A1")
    assert info["value"] == "item"
    assert info["formula"] is None
    for key in ("value", "formula", "data_type", "number_format", "font", "fill_color", "is_merged", "comment"):
        assert key in info


def test_get_cell_merged(tmp_path):
    path = tmp_path / "merged.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "x"
    ws.merge_cells("A1:B1")
    wb.save(path)

    info = reader.get_cell(str(path), "Data", "A1")
    assert info["is_merged"] is True


def test_workbook_info(workbook_path):
    info = reader.workbook_info(workbook_path)
    assert "Data" in [s["name"] for s in info["sheets"]]
    for key in ("name", "dimensions", "max_row", "max_column", "sheet_state"):
        assert key in info["sheets"][0]
    assert info["active_sheet"] == "Data"
    assert isinstance(info["defined_names"], list)


def test_search_workbook_finds_match(workbook_path):
    matches = reader.search_workbook(workbook_path, "apple")
    assert matches == [{"sheet": "Data", "cell": "A2", "value": "apple"}]


def test_search_workbook_limit(workbook_path):
    matches = reader.search_workbook(workbook_path, "", limit=1)
    assert len(matches) == 1


def test_get_cell_formula_value(tmp_path):
    path = tmp_path / "formula.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "=1+1"
    wb.save(path)

    info = reader.get_cell(str(path), "Data", "A1")
    assert info["formula"] == "=1+1"
    assert "value" in info


def test_search_workbook_missing_sheet_raises(workbook_path):
    with pytest.raises(SheetNotFoundError):
        reader.search_workbook(workbook_path, "apple", sheet="NoSuchSheet")
