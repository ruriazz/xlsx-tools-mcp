import openpyxl
import pytest

from xlsx_tools_mcp.errors import SheetNotFoundError
from xlsx_tools_mcp.io import reader, writer


@pytest.fixture
def workbook_path(tmp_path):
    path = tmp_path / "book.xlsx"
    return str(writer.create_workbook(str(path), sheets=["Data"])["path"])


def test_create_workbook_refuses_overwrite_by_default(tmp_path):
    path = str(tmp_path / "book.xlsx")
    writer.create_workbook(path)
    with pytest.raises(FileExistsError):
        writer.create_workbook(path)


def test_create_workbook_refuses_non_excel_path(tmp_path):
    with pytest.raises(ValueError):
        writer.create_workbook(str(tmp_path / "data.txt"))


def test_create_workbook_response_shape(tmp_path):
    result = writer.create_workbook(str(tmp_path / "new.xlsx"), sheets=["A"])
    for key in ("saved", "recalculated", "errors_found", "message", "path", "sheets"):
        assert key in result
    assert result["recalculated"] is False
    assert result["errors_found"] == []


def test_write_cells_then_read_sheet_roundtrip(workbook_path):
    writer.write_cells(workbook_path, "Data", [{"cell": "A1", "value": "hello"}])
    result = reader.read_sheet(workbook_path, "Data")
    assert result["rows"][0][0] == "hello"


def test_write_cells_missing_cell_key_raises(workbook_path):
    with pytest.raises(ValueError):
        writer.write_cells(workbook_path, "Data", [{"value": "oops"}])


def test_write_cells_missing_sheet_raises(workbook_path):
    with pytest.raises(SheetNotFoundError):
        writer.write_cells(workbook_path, "NoSuchSheet", [{"cell": "A1", "value": 1}])


def test_write_cells_survives_process_crash_mid_save(workbook_path, monkeypatch):
    """A crash between the temp-file save and the atomic rename must never
    leave the original file missing or truncated (see `_atomic_save`)."""
    original_save = openpyxl.Workbook.save

    def boom(self, path):
        original_save(self, path)
        raise RuntimeError("simulated crash after temp file write")

    monkeypatch.setattr(openpyxl.Workbook, "save", boom)

    with pytest.raises(RuntimeError):
        writer.write_cells(workbook_path, "Data", [{"cell": "A1", "value": "should not apply"}])

    # Original file must still be intact and readable — crash happened only on the temp copy.
    result = reader.read_sheet(workbook_path, "Data")
    assert result["rows"] == []


def test_delete_sheet_refuses_last_sheet(workbook_path):
    with pytest.raises(ValueError):
        writer.delete_sheet(workbook_path, "Data")


def test_merge_cells(workbook_path):
    result = writer.merge_cells(workbook_path, "Data", "A1:B1")
    assert result["saved"] is True


def test_write_cells_empty_is_noop(workbook_path):
    result = writer.write_cells(workbook_path, "Data", [])
    assert result == {
        "saved": True,
        "recalculated": False,
        "errors_found": [],
        "message": "Nothing to write; file unchanged.",
    }
    assert reader.read_sheet(workbook_path, "Data")["rows"] == []


def test_append_rows_empty_is_noop(workbook_path):
    result = writer.append_rows(workbook_path, "Data", [])
    assert result["saved"] is True
    assert result["recalculated"] is False
    assert reader.read_sheet(workbook_path, "Data")["rows"] == []


def test_write_cells_recalculate_false(workbook_path):
    result = writer.write_cells(workbook_path, "Data", [{"cell": "A1", "value": 42}], recalculate=False)
    assert result["saved"] is True
    assert result["recalculated"] is False


def test_insert_then_delete_rows(workbook_path):
    writer.write_cells(workbook_path, "Data", [{"cell": "A1", "value": "x"}], recalculate=False)
    result = writer.insert_rows(workbook_path, "Data", 2, count=1, recalculate=False)
    assert result["saved"] is True
    result = writer.delete_rows(workbook_path, "Data", 2, count=1, recalculate=False)
    assert result["saved"] is True
    assert reader.read_sheet(workbook_path, "Data")["rows"][0][0] == "x"


def test_insert_then_delete_columns(workbook_path):
    writer.write_cells(workbook_path, "Data", [{"cell": "A1", "value": "x"}], recalculate=False)
    writer.insert_columns(workbook_path, "Data", 2, count=1, recalculate=False)
    result = writer.delete_columns(workbook_path, "Data", 2, count=1, recalculate=False)
    assert result["saved"] is True
    assert reader.read_sheet(workbook_path, "Data")["rows"][0][0] == "x"


def test_merge_then_unmerge(workbook_path):
    writer.merge_cells(workbook_path, "Data", "A1:B1")
    wb = openpyxl.load_workbook(workbook_path)
    assert "A1:B1" in [str(r) for r in wb["Data"].merged_cells.ranges]
    wb.close()

    writer.unmerge_cells(workbook_path, "Data", "A1:B1")
    wb = openpyxl.load_workbook(workbook_path)
    assert len(wb["Data"].merged_cells.ranges) == 0
    wb.close()


def test_set_cell_style_applies_bold_and_fill(workbook_path):
    writer.write_cells(workbook_path, "Data", [{"cell": "A1", "value": "hi"}], recalculate=False)
    writer.set_cell_style(workbook_path, "Data", "A1", {"bold": True, "bg_color": "FF0000"})

    wb = openpyxl.load_workbook(workbook_path)
    cell = wb["Data"]["A1"]
    assert cell.font.bold is True
    assert cell.fill.start_color.rgb.endswith("FF0000")
    wb.close()
